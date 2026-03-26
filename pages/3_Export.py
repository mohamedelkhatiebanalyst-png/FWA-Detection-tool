"""
pages/3_Export.py — Export
---------------------------
Download flagged or full results as Excel with clean, human-readable column names.
"""

import io

import pandas as pd
import streamlit as st

from config import require_results

st.set_page_config(
    page_title="FWA Detection — Export",
    page_icon="📥",
    layout="wide",
)

require_results(st)

df_result = st.session_state["df_result"]
suspicious_count = int(df_result["is_suspicious"].sum())

# ---------------------------------------------------------------------------
# COLUMN RENAME MAP — internal name → human-readable export name
# ---------------------------------------------------------------------------
EXPORT_COLUMNS = {
    "member_id":                      "Member ID",
    "total_claimed":                  "Total Claimed",
    "total_accepted":                 "Total Approved",
    "overbilling_ratio":              "Claim/Approval Ratio",
    "avg_claimed":                    "Avg Claimed per Visit",
    "visit_count":                    "Visit Count",
    "distinct_providers":             "Providers Visited",
    "max_provider_overbilling_ratio": "Worst Provider Ratio",
    "max_doctors_at_prov":            "Max Doctors at One Clinic",
    "max_providers_per_icd":          "Max Providers per Diagnosis",
    "max_same_type_same_day":         "Max Same-Type Claims per Day",
    "inpatient_count":                "Hospital Admissions",
    "distinct_claim_types":           "Claim Type Variety",
    "chronic_ratio":                  "Chronic Claims Ratio",
    "l1_high_freq":                   "Flag: High Visit Frequency",
    "l1_high_cost":                   "Flag: High Total Claimed",
    "l1_high_avg_cost":               "Flag: High Avg Cost per Visit",
    "l2_burst_claims":                "Flag: Repeated Non-Doctor Claims",
    "l2_same_day_multi_prov":         "Flag: Same-Day Multi-Provider",
    "l2_multi_city_same_day":         "Flag: Multi-City Same Day",
    "l2_high_monthly_velocity":       "Flag: High Monthly Volume",
    "l3_provider_overbilling":        "Flag: Provider Overbilling",
    "l3_provider_high_rejection":     "Flag: High Rejection Rate",
    "l4_provider_shopping":           "Flag: Provider Shopping",
    "l4_doctor_mill":                 "Flag: High Doctor Rotation",
    "l4_icd_shopping":                "Flag: Diagnosis Shopping",
    "l5_duplicate_service_same_day":  "Flag: Duplicate Service Same Day",
    "l5_inpatient_frequency":         "Flag: High Inpatient Frequency",
    "l5_claim_type_diversity":        "Flag: High Claim Type Variety",
    "l5_chronic_high_ratio":          "Flag: High Chronic Ratio",
    "fraud_score":                    "Fraud Score",
    "risk_level":                     "Risk Level",
    "is_suspicious":                  "Suspicious",
    "reason":                         "Triggered Rules",
}

RISK_COLORS = {"High": "FFC7CE", "Medium": "FFEB9C", "Low": "C6EFCE"}


def _to_excel(df: pd.DataFrame) -> bytes:
    renamed = df.rename(columns={k: v for k, v in EXPORT_COLUMNS.items() if k in df.columns})
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        renamed.to_excel(writer, index=False, sheet_name="FWA Results")
        ws = writer.sheets["FWA Results"]

        # Freeze header row
        ws.freeze_panes = "A2"

        # Bold + light-grey fill on header
        from openpyxl.styles import Alignment, Font, PatternFill
        header_fill = PatternFill("solid", fgColor="D9D9D9")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Color-code Risk Level column
        risk_col_name = EXPORT_COLUMNS.get("risk_level", "Risk Level")
        risk_col_idx = next(
            (i for i, cell in enumerate(ws[1], 1) if cell.value == risk_col_name), None
        )
        if risk_col_idx:
            for row in ws.iter_rows(min_row=2, min_col=risk_col_idx, max_col=risk_col_idx):
                for cell in row:
                    color = RISK_COLORS.get(cell.value)
                    if color:
                        cell.fill = PatternFill("solid", fgColor=color)

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    return buf.getvalue()


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------
st.title("📥 Export Results")
st.divider()

col1, col2 = st.columns(2)

with col1:
    st.markdown("**Suspicious Members Only**")
    st.caption(f"{suspicious_count:,} members flagged as suspicious.")
    st.download_button(
        label="⬇️ Download Suspicious Members",
        data=_to_excel(df_result[df_result["is_suspicious"]]),
        file_name="fwa_suspicious_members.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

with col2:
    st.markdown("**Full Results (All Members)**")
    st.caption(f"{len(df_result):,} total members with all scores and flags.")
    st.download_button(
        label="⬇️ Download Full Results",
        data=_to_excel(df_result),
        file_name="fwa_all_members.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
